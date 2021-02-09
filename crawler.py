"""
This was is designed to scrape icd10 codes and send them to OE
"""

import os
import time
import random
import unittest
import re
import openpyxl
import platform
import logging
import requests
import statistics
import glob
import json
from lxml.html import fromstring
from datetime import datetime, timedelta
from selenium import webdriver
from fake_useragent import UserAgent
from pathlib import Path
from standalone_tools import eventlog
from bs4 import BeautifulSoup
from selenium.webdriver.remote.remote_connection import LOGGER

class Crawler(object):
    def __init__(self, job_name, timestamp=None):
        """
        Must assign job_name, this is something you can make up.
        May assign timestamp to resume previous job, leave None to scrap whole site again or explicitly set timestamp.
        """
        self.job_name = job_name

        # detect an incomplete database for this job_name otherwise create a new one.
        self.resume, self.timestamp = self.resume_or_create_database()

        self.state = "incomplete"
        
        self.save_crawler_program_data()

        self.webpage_load_interval = 2
        self.driver = None
        self.home_url = None
        self.links_icd10cm_codes = []
        self.links_visited = []
        self.diagnosis_codes = []
        self.measured_seconds_between_completed_pages = [3,3]
        

        
        self.path_job_base = self.make_get_job_paths(self.job_name)
        fp_all_diagnosis_codes_xlsx = os.path.join(self.path_job_base, "output", "All_diagnosis_codes.xlsx")
        
        if not os.path.exists(fp_all_diagnosis_codes_xlsx):
            # we have no xlsx to resume, create a fresh one.
            # xlsx setup
            self.wb_out = openpyxl.Workbook()
            self.ws_out = self.wb_out.active

            self.ws_out["A1"] = "Code"
            self.ws_out["A1"].fill = openpyxl.styles.PatternFill("solid", fgColor="0000CCFF")
            self.ws_out.column_dimensions['A'].width = 15

            self.ws_out["B1"] = "Description"
            self.ws_out["B1"].fill = openpyxl.styles.PatternFill("solid", fgColor="0000CCFF")
            self.ws_out.column_dimensions['B'].width = 100

            self.ws_out["C1"] = "URL"
            self.ws_out["C1"].fill = openpyxl.styles.PatternFill("solid", fgColor="0000CCFF")
            self.ws_out.column_dimensions['C'].width = 120
            self.ws_index = 2
        else:
            # lets resume from the previous xlsx spreadsheet.
            self.wb_out = openpyxl.load_workbook(filename=os.path.join(fp_all_diagnosis_codes_xlsx))
            self.ws_out = self.wb_out.active
            self.ws_index = self.ws_out.max_row + 2

    def _get_timestamp(self) -> str:
        """
        Gets a timestamp in the correct format that is current UTC.
        returns timestamp
        """
        time_string = float(datetime.utcnow().strftime('%y%m%d%H%M%S.%f'))
        time_string = str(round(time_string, 3))
        time_string = time_string.replace(".", "")
        while len(time_string) < 15:
            time_string += "0"
        return time_string

    # for testing a single website - grabbing the source data
    def generate_pretty_source(self, url):
        self.home_url = url
        self.driver = self.make_web_browser()
        self.crawl_url(self.home_url)
        self.shutdown()

    # for downloading the entire site
    def generate_icd10data(self):
        # this must be -1 to start because we need to adjust for the homepage.
        # the homepage is not saved in links_icd10cm codes.
        prev_length_links_icd10cm_codes = -1
        self.driver = self.make_web_browser()
        # if we're not resuming a previous job
        if not self.resume:
            eventlog("===================================")
            eventlog(f"NEW {self.job_name} TIME: {self.timestamp}")
            eventlog("===================================")
            start_time = time.time()
            self.home_url = "https://www.icd10data.com/ICD10CM/Codes"
            self.crawl_url(self.home_url)
            end_time = time.time()
            self.measured_seconds_between_completed_pages.append(end_time - start_time)
        else:
            eventlog("===================================")
            eventlog(f"RESUMING {self.job_name} TIME: {self.timestamp}")
            eventlog("===================================")
            path_job_base = self.make_get_job_paths(self.job_name)
            # we are resuming previous job
            # updated links_visited
            fp_all_links_visited = os.path.join(path_job_base, "output", "links_visited.txt")
            if os.path.exists(fp_all_links_visited):
                with open(fp_all_links_visited, "r", encoding="utf-8") as f:
                    self.links_visited = f.read().splitlines()
                print(f"Total links visited  : {len(self.links_visited)}.")
                print(f"Last link visited was: {self.links_visited[-1]}.")
            # update links_icd10cm_codes
            fp_all_links_icd10cm_codes = os.path.join(path_job_base, "output", "links_icd10cm.txt")
            if os.path.exists(fp_all_links_icd10cm_codes):
                with open(fp_all_links_icd10cm_codes, "r", encoding="utf-8") as f:
                    self.links_icd10cm_codes = f.read().splitlines()
                print(f"Total links_icd10cm_codes: {len(self.links_icd10cm_codes)}.")

        # the + 1 denotes the home page
        # so we visit the home page + all the other links that were found then we are done.
        while len(self.links_visited) != prev_length_links_icd10cm_codes + 1:
            # lets do a depth first search on this website so we can see the data in the spreadsheet sooner
            longest_to_shortest_links = sorted(self.links_icd10cm_codes, key=len, reverse=True)
            eventlog("===================================")
            eventlog("Searching for novel link")
            for url in longest_to_shortest_links:
                start_time = time.time()
                eventlog(f"visiting: {url}")
                found = False

                for visited_link in self.links_visited:
                    if str(visited_link) == str(url):
                        found = True
                        break

                # if current url is not found in visited links, then we havent crawled it.
                if not found:
                    eventlog("Found novel link")
                    self.crawl_url(url)

                    prev_length_links_icd10cm_codes = len(self.links_icd10cm_codes)
                    
                    eventlog("===================================")
                    eventlog(f"{len(self.links_visited)} of {prev_length_links_icd10cm_codes} links visited...")
                    seconds = len(self.links_visited) * statistics.mean(self.measured_seconds_between_completed_pages)
                    minutes = seconds / 60
                    hours = minutes / 60
                    days = hours / 24
                    days_completed = days
                    eventlog("===================================")
                    seconds = prev_length_links_icd10cm_codes * statistics.mean(self.measured_seconds_between_completed_pages)
                    minutes = seconds / 60
                    hours = minutes / 60
                    days = hours / 24
                    days_left = days - days_completed
                    eventlog(f"PROGRESS ESTIMATE: {days_left} days remain...")
                    percentage_completed = (len(self.links_visited) / prev_length_links_icd10cm_codes) * 100
                    eventlog(f"PERCENT COMPLETE : {percentage_completed}")
                    eventlog("===================================")
                end_time = time.time()
                self.measured_seconds_between_completed_pages.append(end_time - start_time)
        
        self.driver.quit()
        self.state = "complete"
        
        self.save_crawler_program_data()


    # detect an incomplete database for this job_name otherwise create a new one.
    def resume_or_create_database(self):
        path_database = str(os.path.join(os.getcwd(), 'DATABASE'))
        if not os.path.exists( path_database ):
            eventlog("Database not found - creating new one.")
            #shutil.rmtree( str(os.getcwd() + '/DATABASE/JOBS/' + str(job_name)))
            os.makedirs( path_database)
            # resume is false, make timestamp.
            return False, self._get_timestamp()
        
        for path in Path(path_database).rglob('crawler_program_data.json'):
            with open(path) as f:
                data = json.load(f)
            
            if str(path).find(self.job_name) != -1:
                if data["state"] != "complete":
                    eventlog(f"Incomplete {self.job_name} resuming timestamp: {data['timestamp']}.")
                    return True, data["timestamp"]

        # resume is false, make timestamp.
        return False, self._get_timestamp()

    def save_crawler_program_data(self):
        data = {
            "state": self.state,
            "timestamp": self.timestamp,
        }

        path_job_base = self.make_get_job_paths(self.job_name)
        fp_crawler_program_data = os.path.join(path_job_base, "output", "crawler_program_data.json")
        with open(fp_crawler_program_data, 'w') as outfile:
            json.dump(data, outfile)

    # downloads url
    # appends it to visited list
    # extracts icd10 code if it exists
    # finds more urls to crawl if they are likely to lead to more icd10 codes 
    def crawl_url(self, url):
        self.links_visited.append(url)
        self.write_link_visited(url)
        self.visit_web_page(url)
        pretty_list = self.download_source_pretty_list(url)

        # find icd code
        self.extract_icd10_code_given_pretty_list(pretty_list)
        # find more links
        self.update_icd10cm_urls(pretty_list)

    def make_web_browser(self):
        LOGGER.setLevel(logging.ERROR)
        logging.getLogger("urllib3").setLevel(logging.ERROR)
        chrome_options = webdriver.ChromeOptions()
        #https://cs.chromium.org/chromium/src/chrome/common/pref_names.cc
        prefs = {
            "profile.managed_default_content_settings.images":2,
            "download.default_directory": "NUL", 
            "download.prompt_for_download": False,
            "download_restrictions":3,
            
        }
        
        chrome_options.accept_untrusted_certs = True
        chrome_options.assume_untrusted_cert_issuer = True
        chrome_options.add_argument('--headless')
        chrome_options.add_argument("--window-size=1400,900")
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument("--disable-impl-side-painting")
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-seccomp-filter-sandbox")
        chrome_options.add_argument("--disable-breakpad")
        chrome_options.add_argument("--disable-client-side-phishing-detection")
        chrome_options.add_argument("--disable-cast")
        chrome_options.add_argument("--disable-cast-streaming-hw-encoding")
        chrome_options.add_argument("--disable-cloud-import")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-session-crashed-bubble")
        chrome_options.add_argument("--disable-ipv6")
        chrome_options.add_argument("--allow-http-screen-capture")
        chrome_options.add_argument("--start-maximized")

        ua = UserAgent()
        userAgent = ua.random
        chrome_options.add_argument(f'user-agent={userAgent}')
        chrome_options.add_experimental_option("prefs", prefs)
        if platform.system() != "Windows":
            # assuming linux - also make sure that the chromedriver version is right before using it on linux
            self.driver = webdriver.Chrome(os.path.join(str(Path.cwd()), 'chromedriver_88'), chrome_options=chrome_options)
        else:
            self.driver = webdriver.Chrome(os.path.join(str(Path.cwd()), 'chromedriver_88.exe'), chrome_options=chrome_options)
        time.sleep(1)
        return self.driver

    # loads web page.
    def visit_web_page(self, url):
        self.driver.get(url)
        time.sleep(self.webpage_load_interval)

    # creates path based on job name in database folder
    def make_get_job_paths(self, name):
        root_job_path = str(os.path.join(os.getcwd(), 'DATABASE', self.timestamp, 'JOBS', str(name)))
        if not os.path.exists( root_job_path ):
            #shutil.rmtree( str(os.getcwd() + '/DATABASE/JOBS/' + str(job_name)))
            os.makedirs( root_job_path)
        output_path = str(os.path.join(os.getcwd(), 'DATABASE', self.timestamp, 'JOBS', str(name), "output"))
        if not os.path.exists( output_path ):
            #shutil.rmtree( str(os.getcwd() + '/DATABASE/JOBS/' + str(job_name)))
            os.makedirs( output_path)
        return root_job_path

    # creates path based on current url and job name in database folder
    def compute_save_directory(self):
        # print("computing save directory")
        # directory_key = None
        # eventlog(f"job_name: {self.job_name}")
        # eventlog(f"job root path: {self.path_job_root}")
        fp_url = ""
        previous_ch = ""
        for ch in str(self.driver.current_url):
            regex = re.compile('[@_!#$%^&*()<>?/\|}{~:;],.')
            if regex.search(ch) == None and ch != ' ':
                if ch == '/' and previous_ch != '/':
                    if platform.system() != "Windows":
                        fp_url += ch
                    else:
                        fp_url += "\\"
                elif ch.isdigit() == True:
                    fp_url += ch
                elif ch.isalpha() == True:
                    fp_url += ch
            previous_ch = ch

        # eventlog(f"fp_url: {fp_url}")
        temp_string = "".join([c for c in str(self.driver.current_url) if c.isalpha() or c.isdigit() or c==' ']).rstrip()
        string = temp_string.replace(" ", "")
        # eventlog(f"string: {string}")

        if len(string) > 60:
            directory_key = str('{:.60}'.format(str(string)))
        else:
            directory_key = string

        # eventlog(f"directory_key: {directory_key}")

        path = os.path.join(self.path_job_base, 'web', str(fp_url))
        if not os.path.exists(path):
            os.makedirs(path)

        # eventlog(f"compute_save_directory path is: {path}")
        return path
        # iresult = 0
        # dp_google_results = os.path.join(website_root_path, "web", fp_url)
        # matrix_name = str(directory_key)

        # pathStart = os.path.join(website_root_path, "web", fp_url)
        # pathEnd = "_index.html"

    # writes source.html into appropriate database directory
    def download_source_pretty_list(self, url):
        raw = self.driver.page_source
        # fp_raw = os.path.join(self.compute_save_directory(), "source.html")

        fp_pretty_list = os.path.join(self.compute_save_directory(), "source_pretty.html")
        # with open(fp_raw, "w+", encoding="utf-8") as f:
        #     f.write(raw)
        self.soup = BeautifulSoup(str(raw), 'html.parser')
        self.prettify_soup = ''
        self.prettify_soup = self.soup.prettify()
        list_pretty = self.prettify_soup.split('\n')
        with open(fp_pretty_list, "w+", encoding="utf-8") as f:
            for line in list_pretty:
                f.write(line + "\n")
        for line in list_pretty:
            if line.find("403 Forbidden") != -1:
                eventlog("detected 403 Forbidden! Waiting 20 minutes and trying again.")
                time.sleep(1200)
                self.driver.quit()
                time.sleep(3)
                self.driver = self.make_web_browser()
                self.visit_web_page(url)
                self.download_source_pretty_list(url)
        return list_pretty
    
    # for testing extraction of links given an url
    def extract_links_given_url(self, url):
        self.home_url = url
        self.driver = self.make_web_browser()
        self.links_visited.append(url)
        self.visit_web_page(url)
        pretty_list = self.download_source_pretty_list(url)
        self.extract_icd10_code_given_pretty_list(pretty_list)
        self.update_icd10cm_urls(pretty_list)
        self.shutdown()

    def update_icd10cm_urls(self, pretty_list):
        for line in pretty_list:
            if line.find("/ICD10CM/Codes/") != -1:
                found = False
                start_index = line.index("/ICD10CM/Codes/")
                icd10cm_end = ""
                end_index = start_index + 60
                while end_index > len(line) -1:
                    end_index -= 1
                extracting = True
                i = start_index
                while extracting and i < end_index:
                    if line[i] != '"':
                        found = True
                        icd10cm_end += line[i]
                    else:
                        extracting = False
                    i += 1
                
                if found and icd10cm_end.find("-") != -1 and icd10cm_end.find("%") == -1:
                    link = f"https://www.icd10data.com{icd10cm_end}"
                    # append to master list of icd 10 code links
                    found = False
                    for saved_link in self.links_icd10cm_codes:
                        if str(link) == str(saved_link):
                            found = True
                            break
                    if not found:
                        self.links_icd10cm_codes.append(link)
                        self.write_icd10cm_code_link_to_job_root(link)

                    # write links to source directory
                    fp_links_icd10cm_codes = os.path.join(self.compute_save_directory(), "source_icd10cm_list.txt")
                    with open(fp_links_icd10cm_codes, "a+", encoding="utf-8") as f:
                        f.write(link + "\n")

    def write_icd10cm_code_link_to_job_root(self, link):
        path_job_base = self.make_get_job_paths(self.job_name)

        dir_all_links_icd10cm_codes = os.path.join(path_job_base, "output")
        if not os.path.exists(dir_all_links_icd10cm_codes):
            os.makedirs(dir_all_links_icd10cm_codes)

        fp_all_links_icd10cm_codes = os.path.join(path_job_base, "output", "links_icd10cm.txt")
        with open(fp_all_links_icd10cm_codes, "a+", encoding="utf-8") as f:
            f.write(link + "\n")

    def write_link_visited(self, url):
        path_job_base = self.make_get_job_paths(self.job_name)

        dir_all_links_visited = os.path.join(path_job_base, "output")
        if not os.path.exists(dir_all_links_visited):
            os.makedirs(dir_all_links_visited)

        fp_all_links_visited = os.path.join(path_job_base, "output", "links_visited.txt")
        with open(fp_all_links_visited, "a+", encoding="utf-8") as f:
            f.write(url + "\n")

    def find_between(self, s, first, last ):
        try:
            start = s.index( first ) + len( first )
            end = s.index( last, start )
            return s[start:end]
        except ValueError:
            return ''

    def extract_icd10_code_given_pretty_list(self, pretty_list):

        for line in pretty_list:
            if line.find("ICD-10-CM Diagnosis Code ") != -1:
                print(f"FOUND DIAGNOSIS CODE: {line} URL: {self.driver.current_url}")
                self.diagnosis_codes.append(line)
                path_job_base = self.make_get_job_paths(self.job_name)
                dir_all_diagnosis_codes = os.path.join(path_job_base, "output")
                if not os.path.exists(dir_all_diagnosis_codes):
                    os.makedirs(dir_all_diagnosis_codes)
                fp_all_diagnosis_codes = os.path.join(path_job_base, "output", "All_diagnosis_codes.txt")
                with open(fp_all_diagnosis_codes, "a+", encoding="utf-8") as f:
                    f.write(line + "\n")

                code = self.find_between(line, "ICD-10-CM Diagnosis Code ", ":")
                description = line.split(f"{code}: ")[-1]

                self.ws_out["A"+str(self.ws_index)] = f"{code}"
                self.ws_out["B"+str(self.ws_index)] = f"{description}"
                self.ws_out["C"+str(self.ws_index)] = f"{self.driver.current_url}"
                
                self.ws_index += 1
                fp_all_diagnosis_codes_xlsx = os.path.join(path_job_base, "output", "All_diagnosis_codes.xlsx")
                self.wb_out.save(fp_all_diagnosis_codes_xlsx)
                break

    # teardown anything we need to
    def shutdown(self):
        self.driver.quit()

class test_unittest(unittest.TestCase):
    
    # # # do entire job
    # this will scann all jobs under the matching job name
    # detect their previous status
    # resume incomplete job
    # start a new job if the latest job was complete.
    def test_crawler(self):
        crawler = Crawler(job_name="live_crawler")
        crawler.generate_icd10data()

    # # pull just pretty source for review.
    # def test_dl_pretty_source(self):
    #     crawler = Crawler("dl_pretty_source")
    #     # crawler.generate_pretty_source("https://www.icd10data.com/ICD10CM/Codes/A00-B99/A00-A09/A00-/A00.0")
    #     crawler.generate_pretty_source("https://www.google.com/")

    # # # pull extract links given filepath
    # def test_extract_links_given_url(self):
    #     crawler = Crawler("dl_pretty_source")
    #     crawler.extract_links_given_url("https://www.icd10data.com/ICD10CM/Codes/Q00-Q99/Q20-Q28")


if __name__ == "__main__":
    # for testing.
    unittest.main()
